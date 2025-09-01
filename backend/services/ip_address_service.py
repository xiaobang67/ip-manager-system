"""
IP地址管理服务
"""
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from models import IPAddress, NetworkSegment, User, Department, IPUsageHistory
from app.schemas import IPAddressCreate, IPAddressUpdate, IPStatusEnum
from fastapi import HTTPException
from datetime import datetime
import ipaddress
import pandas as pd
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class IPAddressService:
    """IP地址管理服务类"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_ip_addresses(self, skip: int = 0, limit: int = 100,
                        network_segment_id: Optional[int] = None,
                        status: Optional[IPStatusEnum] = None,
                        assigned_user_id: Optional[int] = None,
                        assigned_department_id: Optional[int] = None,
                        search: Optional[str] = None) -> List[IPAddress]:
        """获取IP地址列表"""
        query = self.db.query(IPAddress)
        
        # 按网段筛选
        if network_segment_id:
            query = query.filter(IPAddress.network_segment_id == network_segment_id)
        
        # 按状态筛选
        if status:
            query = query.filter(IPAddress.status == status)
        
        # 按分配用户筛选
        if assigned_user_id:
            query = query.filter(IPAddress.assigned_user_id == assigned_user_id)
        
        # 按分配部门筛选
        if assigned_department_id:
            query = query.filter(IPAddress.assigned_department_id == assigned_department_id)
        
        # 搜索功能
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    IPAddress.ip_address.like(search_pattern),
                    IPAddress.device_name.like(search_pattern),
                    IPAddress.hostname.like(search_pattern),
                    IPAddress.mac_address.like(search_pattern),
                    IPAddress.purpose.like(search_pattern)
                )
            )
        
        return query.offset(skip).limit(limit).all()
    
    def get_ip_addresses_with_pagination(self, skip: int = 0, limit: int = 100,
                        network_segment_id: Optional[int] = None,
                        status: Optional[IPStatusEnum] = None,
                        assigned_user_id: Optional[int] = None,
                        assigned_department_id: Optional[int] = None,
                        search: Optional[str] = None) -> tuple[List[IPAddress], int]:
        """获取IP地址列表和总数（用于分页）"""
        query = self.db.query(IPAddress)
        
        # 按网段筛选
        if network_segment_id:
            query = query.filter(IPAddress.network_segment_id == network_segment_id)
        
        # 按状态筛选
        if status:
            query = query.filter(IPAddress.status == status)
        
        # 按分配用户筛选
        if assigned_user_id:
            query = query.filter(IPAddress.assigned_user_id == assigned_user_id)
        
        # 按分配部门筛选
        if assigned_department_id:
            query = query.filter(IPAddress.assigned_department_id == assigned_department_id)
        
        # 搜索功能
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    IPAddress.ip_address.like(search_pattern),
                    IPAddress.device_name.like(search_pattern),
                    IPAddress.hostname.like(search_pattern),
                    IPAddress.mac_address.like(search_pattern),
                    IPAddress.purpose.like(search_pattern)
                )
            )
        
        # 获取总数
        total = query.count()
        
        # 获取分页数据
        items = query.offset(skip).limit(limit).all()
        
        return items, total
    
    def get_ip_address_by_id(self, ip_id: int) -> Optional[IPAddress]:
        """根据ID获取IP地址"""
        return self.db.query(IPAddress).filter(IPAddress.id == ip_id).first()
    
    def get_ip_address_by_ip(self, ip_address: str) -> Optional[IPAddress]:
        """根据IP地址获取记录"""
        return self.db.query(IPAddress).filter(IPAddress.ip_address == ip_address).first()
    
    def create_ip_address(self, ip_data: IPAddressCreate) -> IPAddress:
        """创建IP地址记录"""
        try:
            # 验证IP地址格式
            try:
                ipaddress.ip_address(ip_data.ip_address)
            except ValueError:
                raise HTTPException(status_code=400, detail="IP地址格式不正确")
            
            # 检查IP地址是否已存在
            existing_ip = self.get_ip_address_by_ip(ip_data.ip_address)
            if existing_ip:
                raise HTTPException(status_code=400, detail="IP地址已存在")
            
            # 检查网段是否存在
            segment = self.db.query(NetworkSegment).filter(NetworkSegment.id == ip_data.network_segment_id).first()
            if not segment:
                raise HTTPException(status_code=400, detail="网段不存在")
            
            # 验证IP地址是否在指定网段范围内
            if not self._is_ip_in_segment(ip_data.ip_address, segment):
                raise HTTPException(status_code=400, detail="IP地址不在指定网段范围内")
            
            # 检查分配用户是否存在
            if ip_data.assigned_user_id:
                user = self.db.query(User).filter(User.id == ip_data.assigned_user_id).first()
                if not user:
                    raise HTTPException(status_code=400, detail="分配用户不存在")
            
            # 检查分配部门是否存在
            if ip_data.assigned_department_id:
                dept = self.db.query(Department).filter(Department.id == ip_data.assigned_department_id).first()
                if not dept:
                    raise HTTPException(status_code=400, detail="分配部门不存在")
            
            # 如果状态是已分配，设置分配时间
            if ip_data.status == IPStatusEnum.allocated:
                ip_data_dict = ip_data.model_dump()
                ip_data_dict['allocated_at'] = datetime.now()
                db_ip = IPAddress(**ip_data_dict)
            else:
                db_ip = IPAddress(**ip_data.model_dump())
            
            self.db.add(db_ip)
            self.db.commit()
            self.db.refresh(db_ip)
            
            # 记录操作历史
            try:
                self._record_usage_history(db_ip, 'allocate', None, db_ip.status)
            except Exception as e:
                # 历史记录失败不影响主流程
                print(f"记录操作历史失败: {str(e)}")
            
            return db_ip
            
        except HTTPException:
            self.db.rollback()
            raise
        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=500, detail=f"创建IP地址记录失败: {str(e)}")
    
    def update_ip_address(self, ip_id: int, ip_data: IPAddressUpdate) -> IPAddress:
        """更新IP地址"""
        db_ip = self.get_ip_address_by_id(ip_id)
        if not db_ip:
            raise HTTPException(status_code=404, detail="IP地址不存在")
        
        old_status = db_ip.status
        
        # 检查分配用户
        if ip_data.assigned_user_id:
            user = self.db.query(User).filter(User.id == ip_data.assigned_user_id).first()
            if not user:
                raise HTTPException(status_code=400, detail="分配用户不存在")
        
        # 检查分配部门
        if ip_data.assigned_department_id:
            dept = self.db.query(Department).filter(Department.id == ip_data.assigned_department_id).first()
            if not dept:
                raise HTTPException(status_code=400, detail="分配部门不存在")
        
        # 状态变更处理
        if ip_data.status and ip_data.status != old_status:
            if ip_data.status == IPStatusEnum.allocated and old_status != IPStatusEnum.allocated:
                # 设置分配时间
                ip_data.allocated_at = datetime.now()
            elif ip_data.status != IPStatusEnum.allocated and old_status == IPStatusEnum.allocated:
                # 清除分配信息
                ip_data.allocated_at = None
                ip_data.assigned_user_id = None
                ip_data.assigned_department_id = None
        
        # 更新最后检测时间
        if ip_data.status == IPStatusEnum.allocated:
            ip_data.last_seen = datetime.now()
        
        # 更新字段
        update_data = ip_data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_ip, field, value)
        
        self.db.commit()
        self.db.refresh(db_ip)
        
        # 记录操作历史
        self._record_usage_history(db_ip, 'modify', old_status, db_ip.status)
        
        return db_ip
    
    def delete_ip_address(self, ip_id: int) -> bool:
        """删除IP地址"""
        db_ip = self.get_ip_address_by_id(ip_id)
        if not db_ip:
            raise HTTPException(status_code=404, detail="IP地址不存在")
        
        # 检查是否为已分配状态
        if db_ip.status == IPStatusEnum.allocated:
            raise HTTPException(status_code=400, detail="已分配的IP地址无法删除，请先释放")
        
        # 记录操作历史
        self._record_usage_history(db_ip, 'release', db_ip.status, 'deleted')
        
        # 删除记录
        self.db.delete(db_ip)
        self.db.commit()
        
        return True
    
    def allocate_ip_address(self, ip_id: int, user_id: Optional[int] = None, 
                          department_id: Optional[int] = None,
                          device_info: Optional[dict] = None) -> IPAddress:
        """分配IP地址"""
        db_ip = self.get_ip_address_by_id(ip_id)
        if not db_ip:
            raise HTTPException(status_code=404, detail="IP地址不存在")
        
        if db_ip.status != IPStatusEnum.available:
            raise HTTPException(status_code=400, detail="IP地址不可分配")
        
        # 检查用户和部门
        if user_id:
            user = self.db.query(User).filter(User.id == user_id).first()
            if not user:
                raise HTTPException(status_code=400, detail="用户不存在")
            db_ip.assigned_user_id = user_id
        
        if department_id:
            dept = self.db.query(Department).filter(Department.id == department_id).first()
            if not dept:
                raise HTTPException(status_code=400, detail="部门不存在")
            db_ip.assigned_department_id = department_id
        
        # 更新设备信息
        if device_info:
            db_ip.device_name = device_info.get('device_name')
            db_ip.device_type = device_info.get('device_type')
            db_ip.mac_address = device_info.get('mac_address')
            db_ip.hostname = device_info.get('hostname')
            db_ip.os_type = device_info.get('os_type')
            db_ip.purpose = device_info.get('purpose')
        
        # 更新状态和时间
        old_status = db_ip.status
        db_ip.status = IPStatusEnum.allocated
        db_ip.allocated_at = datetime.now()
        db_ip.last_seen = datetime.now()
        
        self.db.commit()
        self.db.refresh(db_ip)
        
        # 记录操作历史
        self._record_usage_history(db_ip, 'allocate', old_status, db_ip.status)
        
        return db_ip
    
    def release_ip_address(self, ip_id: int) -> IPAddress:
        """释放IP地址"""
        db_ip = self.get_ip_address_by_id(ip_id)
        if not db_ip:
            raise HTTPException(status_code=404, detail="IP地址不存在")
        
        if db_ip.status != IPStatusEnum.allocated:
            raise HTTPException(status_code=400, detail="IP地址未分配，无需释放")
        
        old_status = db_ip.status
        
        # 清除分配信息
        db_ip.status = IPStatusEnum.available
        db_ip.assigned_user_id = None
        db_ip.assigned_department_id = None
        db_ip.device_name = None
        db_ip.device_type = None
        db_ip.mac_address = None
        db_ip.hostname = None
        db_ip.os_type = None
        db_ip.purpose = None
        db_ip.notes = None
        db_ip.allocated_at = None
        db_ip.expires_at = None
        
        self.db.commit()
        self.db.refresh(db_ip)
        
        # 记录操作历史
        self._record_usage_history(db_ip, 'release', old_status, db_ip.status)
        
        return db_ip
    
    def get_ip_usage_history(self, ip_address: str) -> List[IPUsageHistory]:
        """获取IP地址使用历史"""
        return self.db.query(IPUsageHistory).filter(
            IPUsageHistory.ip_address == ip_address
        ).order_by(IPUsageHistory.operation_time.desc()).all()
    
    def batch_allocate_ips(self, segment_id: int, count: int, 
                          user_id: Optional[int] = None,
                          department_id: Optional[int] = None) -> List[IPAddress]:
        """批量分配IP地址"""
        # 获取可用IP地址
        available_ips = self.db.query(IPAddress).filter(
            and_(
                IPAddress.network_segment_id == segment_id,
                IPAddress.status == IPStatusEnum.available
            )
        ).limit(count).all()
        
        if len(available_ips) < count:
            raise HTTPException(status_code=400, detail=f"可用IP地址不足，只有{len(available_ips)}个")
        
        allocated_ips = []
        for ip in available_ips:
            try:
                allocated_ip = self.allocate_ip_address(ip.id, user_id, department_id)
                allocated_ips.append(allocated_ip)
            except Exception as e:
                # 回滚已分配的IP
                for allocated in allocated_ips:
                    self.release_ip_address(allocated.id)
                raise HTTPException(status_code=500, detail=f"批量分配失败: {str(e)}")
        
        return allocated_ips
    
    def _is_ip_in_segment(self, ip_address: str, segment: NetworkSegment) -> bool:
        """检查IP地址是否在网段范围内"""
        try:
            ip = ipaddress.ip_address(ip_address)
            start_ip = ipaddress.ip_address(segment.start_ip)
            end_ip = ipaddress.ip_address(segment.end_ip)
            return start_ip <= ip <= end_ip
        except ValueError:
            return False
    
    def _record_usage_history(self, ip: IPAddress, action: str, old_status: str, new_status: str):
        """记录IP使用历史"""
        try:
            history = IPUsageHistory(
                ip_address=ip.ip_address,
                action=action,
                old_status=old_status,
                new_status=new_status,
                user_id=ip.assigned_user_id,
                department_id=ip.assigned_department_id,
                device_info={
                    'device_name': ip.device_name,
                    'device_type': ip.device_type,
                    'mac_address': ip.mac_address,
                    'hostname': ip.hostname,
                    'os_type': ip.os_type
                } if ip.device_name else None,
                notes=f"IP地址{action}操作"
            )
            self.db.add(history)
            self.db.commit()
        except Exception as e:
            print(f"记录历史失败: {str(e)}")
    
    def batch_create_ips(self, ip_list: List[IPAddressCreate]) -> List[IPAddress]:
        """批量创建IP地址"""
        created_ips = []
        errors = []
        
        for i, ip_data in enumerate(ip_list):
            try:
                created_ip = self.create_ip_address(ip_data)
                created_ips.append(created_ip)
            except Exception as e:
                errors.append(f"第{i+1}行: {str(e)}")
                continue
        
        if errors:
            # 如果有错误，但也有成功的，返回部分成功的结果
            print(f"批量创建部分失败: {errors}")
        
        return created_ips
    
    def batch_import_from_excel(self, file_content: bytes) -> dict:
        """从Excel文件批量导入IP地址"""
        try:
            import pandas as pd
            import io
            
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content))
            
            # 检查必需列
            required_columns = ['ip_address', 'network_segment_id', 'status']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise HTTPException(
                    status_code=400, 
                    detail=f"缺少必需列: {', '.join(missing_columns)}"
                )
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # 构建IP创建数据
                    ip_data = IPAddressCreate(
                        ip_address=str(row['ip_address']),
                        network_segment_id=int(row['network_segment_id']),
                        status=row.get('status', 'available'),
                        device_name=row.get('device_name') if pd.notna(row.get('device_name')) else None,
                        device_type=row.get('device_type') if pd.notna(row.get('device_type')) else None,
                        mac_address=row.get('mac_address') if pd.notna(row.get('mac_address')) else None,
                        hostname=row.get('hostname') if pd.notna(row.get('hostname')) else None,
                        os_type=row.get('os_type') if pd.notna(row.get('os_type')) else None,
                        purpose=row.get('purpose') if pd.notna(row.get('purpose')) else None,
                        notes=row.get('notes') if pd.notna(row.get('notes')) else None,
                        assigned_user_id=int(row['assigned_user_id']) if pd.notna(row.get('assigned_user_id')) and str(row.get('assigned_user_id')).strip() != '' else None,
                        assigned_department_id=int(row['assigned_department_id']) if pd.notna(row.get('assigned_department_id')) and str(row.get('assigned_department_id')).strip() != '' else None
                    )
                    
                    # 创建IP地址
                    self.create_ip_address(ip_data)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"第{index+2}行: {str(e)}")
                    continue
            
            return {
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors
            }
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"解析Excel文件失败: {str(e)}")
    
    def export_to_excel(self, network_segment_id: Optional[int] = None,
                       status: Optional[IPStatusEnum] = None,
                       assigned_user_id: Optional[int] = None,
                       assigned_department_id: Optional[int] = None,
                       search: Optional[str] = None) -> bytes:
        """导出IP地址列表到Excel"""
        try:
            import pandas as pd
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 获取IP地址数据
            ips = self.get_ip_addresses(
                skip=0,
                limit=10000,  # 设置一个较大的限制
                network_segment_id=network_segment_id,
                status=status,
                assigned_user_id=assigned_user_id,
                assigned_department_id=assigned_department_id,
                search=search
            )
            
            # 准备数据
            data = []
            for ip in ips:
                data.append({
                    'IP地址': ip.ip_address,
                    '状态': self._get_status_text(ip.status),
                    '设备名称': ip.device_name or '',
                    '主机名': ip.hostname or '',
                    'MAC地址': ip.mac_address or '',
                    '分配用户': ip.assigned_user.real_name if hasattr(ip, 'assigned_user') and ip.assigned_user else '',
                    '分配部门': ip.assigned_department.name if hasattr(ip, 'assigned_department') and ip.assigned_department else '',
                    '网段': ip.network_segment.name if hasattr(ip, 'network_segment') and ip.network_segment else '',
                    '设备类型': ip.device_type or '',
                    '操作系统': ip.os_type or '',
                    '用途说明': ip.purpose or '',
                    '分配时间': ip.allocated_at.strftime('%Y-%m-%d %H:%M:%S') if ip.allocated_at else '',
                    '最后检测': ip.last_seen.strftime('%Y-%m-%d %H:%M:%S') if ip.last_seen else '',
                    '备注': ip.notes or ''
                })
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "IP地址列表"
            
            # 设置标题样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 设置标题样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 调整列宽
            column_widths = {
                'A': 15,  # IP地址
                'B': 10,  # 状态
                'C': 15,  # 设备名称
                'D': 15,  # 主机名
                'E': 18,  # MAC地址
                'F': 12,  # 分配用户
                'G': 12,  # 分配部门
                'H': 15,  # 网段
                'I': 12,  # 设备类型
                'J': 12,  # 操作系统
                'K': 20,  # 用途说明
                'L': 18,  # 分配时间
                'M': 18,  # 最后检测
                'N': 20   # 备注
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 保存到内存
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"导出Excel失败: {str(e)}")
    
    def _get_status_text(self, status: str) -> str:
        """获取状态中文文本"""
        status_map = {
            'available': '可用',
            'allocated': '已分配',
            'reserved': '保留',
            'blacklisted': '黑名单'
        }
        return status_map.get(status, status)