"""
IP地址管理API路由
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from database.connection import get_db
from services.ip_address_service import IPAddressService
from app.schemas import (
    IPAddressCreate, IPAddressUpdate, IPAddressResponse,
    MessageResponse, IPStatusEnum, OSTypeEnum, PaginatedResponse
)
import io

router = APIRouter()


@router.get("/", response_model=PaginatedResponse[IPAddressResponse])
async def get_ip_addresses(
    skip: int = Query(0, ge=0, description="跳过的记录数"),
    limit: int = Query(100, ge=1, le=1000, description="每页记录数"),
    network_segment_id: Optional[int] = Query(None, description="网段ID"),
    status: Optional[IPStatusEnum] = Query(None, description="IP状态"),
    assigned_user_id: Optional[int] = Query(None, description="分配用户ID"),
    assigned_department_id: Optional[int] = Query(None, description="分配部门ID"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    db: Session = Depends(get_db)
):
    """获取IP地址列表"""
    service = IPAddressService(db)
    # 获取IP地址列表和总数
    ips, total = service.get_ip_addresses_with_pagination(
        skip=skip,
        limit=limit,
        network_segment_id=network_segment_id,
        status=status,
        assigned_user_id=assigned_user_id,
        assigned_department_id=assigned_department_id,
        search=search
    )
    
    # 计算分页信息
    page = skip // limit + 1
    pages = (total + limit - 1) // limit
    
    # 返回分页响应
    return {
        "items": ips,
        "total": total,
        "page": page,
        "size": limit,
        "pages": pages
    }


@router.get("/options/os-types")
async def get_os_types():
    """获取操作系统类型选项"""
    return [
        {"label": "Windows Server 2019", "value": "Windows Server 2019"},
        {"label": "Windows Server 2022", "value": "Windows Server 2022"},
        {"label": "Windows 10", "value": "Windows 10"},
        {"label": "Windows 11", "value": "Windows 11"},
        {"label": "Ubuntu 20.04", "value": "Ubuntu 20.04"},
        {"label": "Ubuntu 22.04", "value": "Ubuntu 22.04"},
        {"label": "CentOS 7", "value": "CentOS 7"},
        {"label": "CentOS 8", "value": "CentOS 8"},
        {"label": "Debian 10", "value": "Debian 10"},
        {"label": "Debian 11", "value": "Debian 11"},
        {"label": "Red Hat Enterprise Linux 8", "value": "Red Hat Enterprise Linux 8"},
        {"label": "Red Hat Enterprise Linux 9", "value": "Red Hat Enterprise Linux 9"},
        {"label": "SUSE Linux Enterprise 15", "value": "SUSE Linux Enterprise 15"},
        {"label": "macOS Monterey", "value": "macOS Monterey"},
        {"label": "macOS Ventura", "value": "macOS Ventura"},
        {"label": "macOS Sonoma", "value": "macOS Sonoma"},
        {"label": "VMware vSphere ESXi 7", "value": "VMware vSphere ESXi 7"},
        {"label": "VMware vSphere ESXi 8", "value": "VMware vSphere ESXi 8"},
        {"label": "其他", "value": "其他"}
    ]


@router.get("/export-excel")
async def export_ips_to_excel(
    network_segment_id: Optional[int] = Query(None, description="网段ID"),
    status: Optional[IPStatusEnum] = Query(None, description="IP状态"),
    assigned_user_id: Optional[int] = Query(None, description="分配用户ID"),
    assigned_department_id: Optional[int] = Query(None, description="分配部门ID"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    db: Session = Depends(get_db)
):
    """导出IP地址列表到Excel"""
    try:
        service = IPAddressService(db)
        excel_buffer = service.export_to_excel(
            network_segment_id=network_segment_id,
            status=status,
            assigned_user_id=assigned_user_id,
            assigned_department_id=assigned_department_id,
            search=search
        )
        
        headers = {
            'Content-Disposition': 'attachment; filename="ip_addresses.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        return StreamingResponse(
            io.BytesIO(excel_buffer),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/batch-import")
async def batch_import_ips(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """批量导入IP地址"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx, .xls)")
    
    try:
        service = IPAddressService(db)
        contents = await file.read()
        result = service.batch_import_from_excel(contents)
        return {
            "message": "导入完成",
            "success_count": result["success_count"],
            "error_count": result["error_count"],
            "errors": result["errors"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.post("/batch-allocate", response_model=List[IPAddressResponse])
async def batch_allocate_ips(
    segment_id: int,
    count: int = Query(..., ge=1, le=100, description="分配数量"),
    user_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """批量分配IP地址"""
    service = IPAddressService(db)
    return service.batch_allocate_ips(segment_id, count, user_id, department_id)


@router.get("/by-ip/{ip_address}", response_model=IPAddressResponse)
async def get_ip_address_by_ip(ip_address: str, db: Session = Depends(get_db)):
    """根据IP地址获取详情"""
    service = IPAddressService(db)
    ip = service.get_ip_address_by_ip(ip_address)
    if not ip:
        raise HTTPException(status_code=404, detail="IP地址不存在")
    return ip


@router.get("/{ip_id}", response_model=IPAddressResponse)
async def get_ip_address(ip_id: int, db: Session = Depends(get_db)):
    """根据ID获取IP地址详情"""
    service = IPAddressService(db)
    ip = service.get_ip_address_by_id(ip_id)
    if not ip:
        raise HTTPException(status_code=404, detail="IP地址不存在")
    return ip


@router.get("/{ip_address}/history")
async def get_ip_usage_history(ip_address: str, db: Session = Depends(get_db)):
    """获取IP地址使用历史"""
    service = IPAddressService(db)
    history = service.get_ip_usage_history(ip_address)
    return {"history": history}


@router.post("/", response_model=IPAddressResponse, status_code=201)
async def create_ip_address(
    ip_data: IPAddressCreate,
    db: Session = Depends(get_db)
):
    """创建IP地址记录"""
    service = IPAddressService(db)
    return service.create_ip_address(ip_data)


@router.put("/{ip_id}", response_model=IPAddressResponse)
async def update_ip_address(
    ip_id: int,
    ip_data: IPAddressUpdate,
    db: Session = Depends(get_db)
):
    """更新IP地址"""
    service = IPAddressService(db)
    return service.update_ip_address(ip_id, ip_data)


@router.delete("/{ip_id}", response_model=MessageResponse)
async def delete_ip_address(ip_id: int, db: Session = Depends(get_db)):
    """删除IP地址"""
    service = IPAddressService(db)
    service.delete_ip_address(ip_id)
    return MessageResponse(message="IP地址删除成功")


@router.post("/{ip_id}/allocate", response_model=IPAddressResponse)
async def allocate_ip_address(
    ip_id: int,
    user_id: Optional[int] = None,
    department_id: Optional[int] = None,
    device_name: Optional[str] = None,
    device_type: Optional[str] = None,
    mac_address: Optional[str] = None,
    hostname: Optional[str] = None,
    os_type: Optional[str] = None,
    purpose: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """分配IP地址"""
    service = IPAddressService(db)
    device_info = {
        'device_name': device_name,
        'device_type': device_type,
        'mac_address': mac_address,
        'hostname': hostname,
        'os_type': os_type,
        'purpose': purpose
    } if any([device_name, device_type, mac_address, hostname, os_type, purpose]) else None
    
    return service.allocate_ip_address(ip_id, user_id, department_id, device_info)


@router.post("/{ip_id}/release", response_model=IPAddressResponse)
async def release_ip_address(ip_id: int, db: Session = Depends(get_db)):
    """释放IP地址"""
    service = IPAddressService(db)
    return service.release_ip_address(ip_id)


@router.post("/batch-create", response_model=List[IPAddressResponse])
async def batch_create_ips(
    ip_list: List[IPAddressCreate],
    db: Session = Depends(get_db)
):
    """批量创建IP地址"""
    service = IPAddressService(db)
    return service.batch_create_ips(ip_list)


@router.get("/download-template")
async def download_excel_template():
    """下载Excel导入模板"""
    try:
        # 创建模板数据
        template_data = {
            'ip_address': ['192.168.1.100', '192.168.1.101'],
            'network_segment_id': [1, 1],
            'status': ['available', 'allocated'],
            'device_name': ['服务妖01', '工作站01'],
            'device_type': ['server', 'workstation'],
            'mac_address': ['00:11:22:33:44:55', '00:11:22:33:44:56'],
            'hostname': ['server01', 'ws01'],
            'os_type': ['Linux', 'Windows 10'],
            'purpose': ['Web服务器', '办公用机'],
            'notes': ['生产环境', '开发环境'],
            'assigned_user_id': ['', '1'],
            'assigned_department_id': ['', '1']
        }
        
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # 创建 DataFrame
        df = pd.DataFrame(template_data)
        
        # 创建 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "IP地址导入模板"
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置标题样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加说明工作表
        ws_info = wb.create_sheet("导入说明")
        info_data = [
            ["字段名", "说明", "是否必需", "示例"],
            ["ip_address", "IP地址", "是", "192.168.1.100"],
            ["network_segment_id", "网段ID", "是", "1"],
            ["status", "状态", "是", "available/allocated/reserved/blacklisted"],
            ["device_name", "设备名称", "否", "服务妖01"],
            ["device_type", "设备类型", "否", "server/workstation/printer/network/other"],
            ["mac_address", "MAC地址", "否", "00:11:22:33:44:55"],
            ["hostname", "主机名", "否", "server01"],
            ["os_type", "操作系统", "否", "Linux/Windows 10"],
            ["purpose", "用途说明", "否", "Web服务器"],
            ["notes", "备注", "否", "生产环境"],
            ["assigned_user_id", "分配用户ID", "否", "1"],
            ["assigned_department_id", "分配部门ID", "否", "1"]
        ]
        
        for row in info_data:
            ws_info.append(row)
        
        # 设置说明表格样式
        for cell in ws_info[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D']:
            ws_info.column_dimensions[col].width = 20
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="ip_import_template.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建模板失败: {str(e)}")