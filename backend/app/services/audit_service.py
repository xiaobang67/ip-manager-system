from typing import List, Optional, Dict, Any, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from app.models.audit_log import AuditLog
from app.models.user import User
from datetime import datetime, timedelta
from app.core.timezone_config import now_beijing
import csv
import json
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# 暂时禁用pandas导入，避免numpy版本兼容性问题
# import pandas as pd


class AuditService:
    def __init__(self, db: Session):
        self.db = db

    def log_operation(
        self,
        user_id: int,
        action: str,
        entity_type: str,
        entity_id: Optional[int] = None,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None
    ) -> AuditLog:
        """记录操作审计日志"""
        audit_log = AuditLog(
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            old_values=old_values,
            new_values=new_values,
            ip_address=ip_address,
            user_agent=user_agent
        )
        
        self.db.add(audit_log)
        self.db.commit()
        self.db.refresh(audit_log)
        
        return audit_log

    def get_entity_history(
        self,
        entity_type: str,
        entity_id: int,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """获取实体的历史记录"""
        logs = (
            self.db.query(AuditLog)
            .filter(
                AuditLog.entity_type == entity_type,
                AuditLog.entity_id == entity_id
            )
            .order_by(desc(AuditLog.created_at))
            .limit(limit)
            .all()
        )
        
        history = []
        for log in logs:
            # 获取用户信息
            user = self.db.query(User).filter(User.id == log.user_id).first()
            username = user.username if user else "Unknown"
            
            history.append({
                "id": log.id,
                "action": log.action,
                "user_id": log.user_id,
                "username": username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "created_at": log.created_at
            })
        
        return history

    def get_user_activity(
        self,
        user_id: int,
        entity_type: Optional[str] = None,
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """获取用户活动记录"""
        query = self.db.query(AuditLog).filter(AuditLog.user_id == user_id)
        
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        
        logs = query.order_by(desc(AuditLog.created_at)).limit(limit).all()
        
        activity = []
        for log in logs:
            activity.append({
                "id": log.id,
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "created_at": log.created_at
            })
        
        return activity

    def get_recent_activities(
        self,
        entity_type: Optional[str] = None,
        action: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """获取最近的活动记录"""
        query = self.db.query(AuditLog)
        
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        
        if action:
            query = query.filter(AuditLog.action == action)
        
        logs = query.order_by(desc(AuditLog.created_at)).limit(limit).all()
        
        activities = []
        for log in logs:
            # 获取用户信息
            user = self.db.query(User).filter(User.id == log.user_id).first()
            username = user.username if user else "Unknown"
            
            activities.append({
                "id": log.id,
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "user_id": log.user_id,
                "username": username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "created_at": log.created_at
            })
        
        return activities

    def search_audit_logs(
        self,
        user_id: Optional[int] = None,
        entity_type: Optional[str] = None,
        action: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        skip: int = 0,
        limit: int = 50
    ) -> List[Dict[str, Any]]:
        """搜索审计日志"""
        query = self.db.query(AuditLog)
        
        if user_id:
            query = query.filter(AuditLog.user_id == user_id)
        
        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)
        
        if action:
            query = query.filter(AuditLog.action == action)
        
        if start_date:
            query = query.filter(AuditLog.created_at >= start_date)
        
        if end_date:
            query = query.filter(AuditLog.created_at <= end_date)
        
        logs = (
            query.order_by(desc(AuditLog.created_at))
            .offset(skip)
            .limit(limit)
            .all()
        )
        
        results = []
        for log in logs:
            # 获取用户信息
            user = self.db.query(User).filter(User.id == log.user_id).first()
            username = user.username if user else "Unknown"
            
            results.append({
                "id": log.id,
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "user_id": log.user_id,
                "username": username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "created_at": log.created_at
            })
        
        return results

    def search_audit_logs_with_count(
        self,
        user_id: Optional[int] = None,
        entity_type: Optional[str] = None,
        action: Optional[str] = None,
        entity_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        skip: int = 0,
        limit: int = 50
    ) -> Tuple[List[Dict[str, Any]], int]:
        """搜索审计日志并返回总数"""
        query = self.db.query(AuditLog)
        
        # 构建过滤条件
        filters = []
        if user_id:
            filters.append(AuditLog.user_id == user_id)
        if entity_type:
            filters.append(AuditLog.entity_type == entity_type)
        if action:
            filters.append(AuditLog.action == action)
        if entity_id:
            filters.append(AuditLog.entity_id == entity_id)
        if start_date:
            filters.append(AuditLog.created_at >= start_date)
        if end_date:
            filters.append(AuditLog.created_at <= end_date)
        
        if filters:
            query = query.filter(and_(*filters))
        
        # 获取总数
        total = query.count()
        
        # 获取分页数据
        logs = (
            query.order_by(desc(AuditLog.created_at))
            .offset(skip)
            .limit(limit)
            .all()
        )
        
        results = []
        for log in logs:
            # 获取用户信息
            user = self.db.query(User).filter(User.id == log.user_id).first()
            username = user.username if user else "Unknown"
            
            results.append({
                "id": log.id,
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "user_id": log.user_id,
                "username": username,
                "old_values": log.old_values,
                "new_values": log.new_values,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "created_at": log.created_at
            })
        
        return results, total

    def get_audit_statistics(self) -> Dict[str, Any]:
        """获取审计日志统计信息"""
        # 总日志数
        total_logs = self.db.query(AuditLog).count()
        
        # 按操作类型统计
        actions_count = {}
        action_stats = (
            self.db.query(AuditLog.action, func.count(AuditLog.id))
            .group_by(AuditLog.action)
            .all()
        )
        for action, count in action_stats:
            actions_count[action] = count
        
        # 按实体类型统计
        entities_count = {}
        entity_stats = (
            self.db.query(AuditLog.entity_type, func.count(AuditLog.id))
            .group_by(AuditLog.entity_type)
            .all()
        )
        for entity_type, count in entity_stats:
            entities_count[entity_type] = count
        
        # 按用户统计
        users_count = {}
        user_stats = (
            self.db.query(AuditLog.user_id, func.count(AuditLog.id))
            .group_by(AuditLog.user_id)
            .all()
        )
        for user_id, count in user_stats:
            user = self.db.query(User).filter(User.id == user_id).first()
            username = user.username if user else f"User_{user_id}"
            users_count[username] = count
        
        # 最近活动
        recent_activities = self.get_recent_activities(limit=10)
        
        return {
            "total_logs": total_logs,
            "actions_count": actions_count,
            "entities_count": entities_count,
            "users_count": users_count,
            "recent_activities": recent_activities
        }

    def export_audit_logs(
        self,
        format_type: str = "csv",
        user_id: Optional[int] = None,
        entity_type: Optional[str] = None,
        action: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        limit: int = 10000
    ) -> bytes:
        """导出审计日志"""
        # 获取数据
        logs, _ = self.search_audit_logs_with_count(
            user_id=user_id,
            entity_type=entity_type,
            action=action,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )
        
        if format_type.lower() == "csv":
            return self._export_to_csv(logs)
        elif format_type.lower() == "excel":
            return self._export_to_excel(logs)
        elif format_type.lower() == "json":
            return self._export_to_json(logs)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")

    def _export_to_csv(self, logs: List[Dict[str, Any]]) -> bytes:
        """导出为CSV格式"""
        output = io.StringIO()
        
        if not logs:
            return b""
        
        fieldnames = [
            "id", "action", "entity_type", "entity_id", "user_id", "username",
            "old_values", "new_values", "ip_address", "user_agent", "created_at"
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for log in logs:
            # 处理JSON字段
            row = log.copy()
            row["old_values"] = json.dumps(log["old_values"]) if log["old_values"] else ""
            row["new_values"] = json.dumps(log["new_values"]) if log["new_values"] else ""
            row["created_at"] = log["created_at"].isoformat() if log["created_at"] else ""
            
            writer.writerow(row)
        
        return output.getvalue().encode('utf-8')

    def _export_to_excel(self, logs: List[Dict[str, Any]]) -> bytes:
        """导出为Excel格式"""
        if not logs:
            # 创建空的Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Logs"
            ws.append(["No data available"])
            
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # 准备数据
        processed_logs = []
        for log in logs:
            row = log.copy()
            row["old_values"] = json.dumps(log["old_values"]) if log["old_values"] else ""
            row["new_values"] = json.dumps(log["new_values"]) if log["new_values"] else ""
            row["created_at"] = log["created_at"].isoformat() if log["created_at"] else ""
            processed_logs.append(row)
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        # 获取列名
        if processed_logs:
            columns = list(processed_logs[0].keys())
            
            # 添加表头
            for col_num, column_title in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)
            
            # 添加数据
            for row_num, log in enumerate(processed_logs, 2):
                for col_num, column in enumerate(columns, 1):
                    value = log.get(column, "")
                    ws.cell(row=row_num, column=col_num, value=str(value))
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _export_to_json(self, logs: List[Dict[str, Any]]) -> bytes:
        """导出为JSON格式"""
        # 处理datetime对象
        processed_logs = []
        for log in logs:
            processed_log = log.copy()
            if processed_log["created_at"]:
                processed_log["created_at"] = processed_log["created_at"].isoformat()
            processed_logs.append(processed_log)
        
        return json.dumps(processed_logs, indent=2, ensure_ascii=False).encode('utf-8')

    def archive_old_logs(self, days_to_keep: int = 365) -> int:
        """归档旧的审计日志"""
        cutoff_date = now_beijing() - timedelta(days=days_to_keep)
        
        # 计算要删除的记录数
        count = (
            self.db.query(AuditLog)
            .filter(AuditLog.created_at < cutoff_date)
            .count()
        )
        
        # 删除旧记录
        self.db.query(AuditLog).filter(AuditLog.created_at < cutoff_date).delete()
        self.db.commit()
        
        return count